# -*- coding: utf-8 -*-
"""Gera o painel web do REA a partir das planilhas do Mateus.

Uso: python gerar_painel.py
Saída: index.html (autocontido, abre em qualquer navegador/celular)
"""
import json
import datetime
import base64
from pathlib import Path

import openpyxl

PASTA_RACHA = Path(r"Z:\DIRETORIA\Mateus\Particular\Racha")
ARQ_RANKING = PASTA_RACHA / "Ranking REA Cópia.xlsx"
ARQ_PAGTO = PASTA_RACHA / "CONTROLE DE PAGTO.xlsx"
ARQ_BALANCETE = PASTA_RACHA / "BALANCETE REA.xlsx"
SAIDA = Path(__file__).parent / "index.html"

MES_PAGTO = "JUL 2026"  # usado só se a aba do mês atual não existir
MES_BALANCETE = "Jul 2026"
# Início do ciclo atual do Craque — o ranking do Craque conta a partir desta data.
CICLO_INICIO = datetime.date(2026, 8, 1)
# Aviso da faixa preta do topo (editar conforme a fase do campeonato)
AVISO_TOPO = ("🏆 1º ciclo: PAGODE campeão! · Torneio 25/07: ROSA campeão! · "
              "🔄 Novo ciclo começou em 01/08 — todos zerados, disputa aberta!")

# Pix do racha (chave aleatória dedicada — não expõe dados pessoais)
PIX_CHAVE = "1716f630-fb4e-45a5-ba26-7629786fcad2"
PIX_NOME = "RACHA ENTRE AMIGOS"
PIX_CIDADE = "SAO JOSE"

# Backend de presença (Google Apps Script). /exec
RSVP_URL = "https://script.google.com/macros/s/AKfycbzqhPslN6qj8MIkwQ18EnvIGNAP3wk9CgUkGwqy203fAjJ11UfO6Ec2-pfn_n9J2FKUXw/exec"

# Elenco para o seletor de presença (atualizar quando entrar/sair atleta)
ROSTER = ["Alexandre", "Armando", "Arnaldo", "Balu", "Bruno", "Cabeça", "Charles",
          "Chiclete", "Douglas", "Egnaldo", "Enzo", "Fabio", "Fernandinho",
          "Filipe", "Francês", "Gabriel", "Hugo", "Iroshi", "Isack", "Katito",
          "Luizinho", "Marcio", "Marlon", "Mateus", "Mauricio", "Pagode", "Pato",
          "Paulinho", "Pena", "Vinicius", "Walter", "Wesley", "Yuri", "Zé"]
MESES_PT = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
            "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]


def _emv(tid, valor):
    return f"{tid}{len(valor):02d}{valor}"


def _crc16(payload):
    crc = 0xFFFF
    for b in payload.encode("utf-8"):
        crc ^= b << 8
        for _ in range(8):
            crc = ((crc << 1) ^ 0x1021) if (crc & 0x8000) else (crc << 1)
            crc &= 0xFFFF
    return f"{crc:04X}"


def pix_copia_cola(chave, nome, cidade, valor=None, txid="***"):
    """Monta o BR Code (Pix copia-e-cola) estático. Valor None = aberto."""
    mai = _emv("00", "br.gov.bcb.pix") + _emv("01", chave)
    p = (
        _emv("00", "01")
        + _emv("26", mai)
        + _emv("52", "0000")
        + _emv("53", "986")
        + (_emv("54", f"{valor:.2f}") if valor else "")
        + _emv("58", "BR")
        + _emv("59", nome[:25])
        + _emv("60", cidade[:15])
        + _emv("62", _emv("05", txid))
    )
    p += "6304"
    return p + _crc16(p)


def pix_qr_base64(payload):
    import io
    import qrcode
    img = qrcode.make(payload)
    buf = io.BytesIO()
    img.save(buf, "PNG")
    return base64.b64encode(buf.getvalue()).decode()


def dados_pix():
    payload = pix_copia_cola(PIX_CHAVE, PIX_NOME, PIX_CIDADE)
    return {"chave": PIX_CHAVE, "copia_cola": payload, "qr": pix_qr_base64(payload)}


def aba_pagto(wb):
    """Aba do mês corrente; se ainda não existir, cai na constante MES_PAGTO."""
    hoje = datetime.date.today()
    nome = f"{MESES_PT[hoje.month - 1]} {hoje.year}"
    return wb[nome] if nome in wb.sheetnames else wb[MES_PAGTO]

# Último racha lançado (atualizar a cada súmula)
ULTIMO_RACHA = {
    "data": "08/08/2026",
    "partidas": [
        "Azul 0 x 0 Amarelo", "Azul 1 x 1 Amarelo", "Azul 0 x 2 Amarelo",
        "Azul 1 x 2 Amarelo", "Azul 1 x 1 Amarelo", "Azul 1 x 4 Amarelo",
        "Azul 3 x 3 Amarelo", "Azul 2 x 2 Amarelo",
    ],
    "quadro": [
        {"time": "Amarelo", "v": 3, "e": 5, "d": 0, "gols": 15},
        {"time": "Azul", "v": 0, "e": 5, "d": 3, "gols": 9},
    ],
    "bola_cheia": "Wesley",
    "bola_murcha": "Cabeça",
}

APELIDOS = {
    "WESLEY IRMÃO": "WESLEY", "WESLEI IRMÃO": "WESLEY", "WESLEI  IRMÃO": "WESLEY",
    "ALEXANDRE GOL": "ALEXANDRE", "EGNALDO GOL": "EGNALDO",
    "IROSHI GOLEIRO": "IROSHI", "FABIO PRESID.": "FABIO",
    "MATEUS TESOUR.": "MATEUS",
}
IGNORAR = {"TOTAL", "TOTAIS", "TOTAIS.......", ""}
SAIRAM = {"EDUARDO"}  # atletas que saíram do racha — não aparecem no site


def nome_norm(v):
    n = " ".join(str(v).split()).upper().strip()
    return APELIDOS.get(n, n)


def somar_aba(ws, apenas_2026=False, desde=None):
    """Soma valores das colunas com data no cabeçalho (linha 2). Retorna {nome: total} e nº de datas presentes.
    desde: se informado (date), soma só os sábados a partir dessa data (usado p/ o ciclo do Craque)."""
    datas = {}
    for c in ws[2]:
        if isinstance(c.value, datetime.datetime):
            if apenas_2026 and c.value.year != 2026:
                continue
            if desde and c.value.date() < desde:
                continue
            datas[c.column] = c.value
    totais, presencas, pontuais = {}, {}, {}
    detalhe = {}
    for r in range(3, ws.max_row + 1):
        raw = ws.cell(r, 2).value
        if raw is None:
            continue
        nome = nome_norm(raw)
        if nome in IGNORAR or nome.startswith("TOTAL"):
            continue
        soma, pres, pont = 0, 0, 0
        det = {}
        for col in datas:
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                soma += v
                pres += 1
                if v == 3:
                    pont += 1
                det[datas[col].strftime("%Y-%m-%d")] = v
        totais[nome] = totais.get(nome, 0) + soma
        presencas[nome] = presencas.get(nome, 0) + pres
        pontuais[nome] = pontuais.get(nome, 0) + pont
        detalhe[nome] = det
    return totais, presencas, pontuais, detalhe, sorted(datas.values())


def ranking(dic, exclui_zero=True):
    itens = [(n, v) for n, v in dic.items() if not (exclui_zero and v == 0)]
    return [
        {"nome": n.title(), "total": (int(v) if v == int(v) else round(v, 1))}
        for n, v in sorted(itens, key=lambda x: -x[1])
    ]


def dados_ranking():
    wb = openpyxl.load_workbook(ARQ_RANKING, data_only=True)
    craque, _, _, _, _ = somar_aba(wb["Craque 2026"], desde=CICLO_INICIO)
    artilharia, _, _, _, _ = somar_aba(wb["Artilharia 2026"])
    vitorias, _, _, _, _ = somar_aba(wb["Vitórias 2026"])
    cartoes, _, _, _, _ = somar_aba(wb["Cartões 2026"])
    bolas, _, _, det_bolas, _ = somar_aba(wb["Bola Cheia 2026"])
    pres_tot, pres_qtd, pres_pont, det_pres, datas_pres = somar_aba(wb["Presença"], apenas_2026=True)

    # bolas: contar troféus (valores 2 = cheia, -1 = murcha)
    cheia, murcha = {}, {}
    for nome, det in det_bolas.items():
        cheia[nome] = sum(1 for v in det.values() if v > 0)
        murcha[nome] = sum(1 for v in det.values() if v < 0)

    # atleta fiel: presente (>=2) nas últimas 8 datas com jogo
    ws = wb["Presença"]
    datas_jogo = []
    for c in ws[2]:
        if isinstance(c.value, datetime.datetime):
            tem_dado = any(
                isinstance(ws.cell(r, c.column).value, (int, float))
                for r in range(3, ws.max_row + 1)
            )
            if tem_dado:
                datas_jogo.append((c.value, c.column))
    datas_jogo.sort()
    ult8 = datas_jogo[-8:]
    fieis, fiel_tab = [], []
    for r in range(3, ws.max_row + 1):
        raw = ws.cell(r, 2).value
        if raw is None:
            continue
        nome = nome_norm(raw)
        if nome in IGNORAR:
            continue
        pontos8 = sum(
            ws.cell(r, col).value
            for _, col in ult8
            if isinstance(ws.cell(r, col).value, (int, float))
        )
        ok = all(
            isinstance(ws.cell(r, col).value, (int, float)) and ws.cell(r, col).value >= 2
            for _, col in ult8
        )
        if ok:
            fieis.append(nome.title())
        fiel_tab.append({"nome": nome.title(), "pontos": int(pontos8), "fiel": ok})
    # fiéis primeiro (agrupados), depois quem não é — cada grupo em ordem alfabética
    fiel_tab.sort(key=lambda x: (not x["fiel"], x["nome"]))

    estat = []
    todos = sorted(set(craque) | set(pres_qtd))
    for n in todos:
        if craque.get(n, 0) == 0 and pres_qtd.get(n, 0) == 0:
            continue
        estat.append({
            "nome": n.title(),
            "vitorias": int(vitorias.get(n, 0)),
            "presencas": int(pres_qtd.get(n, 0)),
            "pontualidade": int(pres_pont.get(n, 0)),
            "cartoes": int(cartoes.get(n, 0)),
            "cheia": int(cheia.get(n, 0)),
            "murcha": int(murcha.get(n, 0)),
            "fiel": n.title() in fieis,
        })
    estat.sort(key=lambda x: -x["presencas"])
    return {
        "craque": ranking(craque),
        "artilharia": ranking(artilharia),
        "estatisticas": estat,
        "fieis": sorted(fieis),
        "fiel_tab": fiel_tab,
    }


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def dados_financeiro():
    wb = openpyxl.load_workbook(ARQ_PAGTO, data_only=True)
    ws = aba_pagto(wb)
    linhas = []
    for r in range(3, 45):
        raw = ws.cell(r, 2).value
        if raw is None:
            continue
        nome = nome_norm(raw)
        if nome in IGNORAR or nome in SAIRAM or nome.startswith("TOTAIS"):
            continue
        sa = num(ws.cell(r, 3).value)  # dívida de meses anteriores
        total = num(ws.cell(r, 6).value)
        pago = num(ws.cell(r, 8).value)
        saldo = num(ws.cell(r, 10).value)
        # dívida antiga que sobrou depois dos pagamentos (pagamento abate o antigo primeiro);
        # só é "Atrasado" quem ainda deve mensalidade de mês anterior (>= 90);
        # multa pequena antiga não rebaixa para atrasado
        restante_antigo = max(0.0, sa - pago)
        if total == 0 and saldo == 0:
            situacao = "Isento"
        elif saldo <= 0:
            situacao = "Em dia"
        elif restante_antigo >= 90:
            situacao = "Atrasado"
        else:
            situacao = "Mês em aberto"
        # obs NÃO vai para o site (nota interna, evita expor no código-fonte público)
        linhas.append({
            "nome": nome.title(), "total": round(total, 2),
            "pago": round(pago, 2), "saldo": round(saldo, 2),
            "situacao": situacao,
        })
    ordem = {"Atrasado": 0, "Mês em aberto": 1, "Em dia": 2, "Isento": 3}
    linhas.sort(key=lambda x: (ordem[x["situacao"]], x["nome"]))
    return linhas


def saldo_caixa():
    wb = openpyxl.load_workbook(ARQ_BALANCETE, data_only=True)
    ws = wb[MES_BALANCETE]
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a and "SALDO ATUAL" in str(a).upper():
            for c in range(ws.max_column, 1, -1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    return round(v, 2)
    return None


def logo_html():
    for nome in ["logo.png", "logo REA.png", "logo.jpg", "logo REA.jpg", "logo.jpeg"]:
        p = Path(__file__).parent / nome
        if not p.exists():
            p = PASTA_RACHA / nome
        if p.exists():
            try:  # reduz a imagem para manter o HTML leve
                import io
                from PIL import Image
                img = Image.open(p)
                img.thumbnail((280, 280))
                buf = io.BytesIO()
                img.save(buf, "PNG", optimize=True)
                dados_img = buf.getvalue()
            except ImportError:
                dados_img = p.read_bytes()
            b64 = base64.b64encode(dados_img).decode()
            return f'<img class="logo" src="data:image/png;base64,{b64}" alt="REA">'
    # fallback: escudo SVG inspirado no logo oficial
    return """<svg class="logo" viewBox="0 0 120 130" xmlns="http://www.w3.org/2000/svg">
      <path d="M60 4 L74 14 L98 10 L96 34 L112 50 L94 66 L96 96 L74 100 L60 122 L46 100 L24 96 L26 66 L8 50 L24 34 L22 10 L46 14 Z"
            fill="#1AA3DD" stroke="#111" stroke-width="6"/>
      <path d="M26 52 Q60 66 94 50 L94 58 Q60 74 26 60 Z" fill="#111"/>
      <path d="M26 66 Q60 80 94 64 L94 72 Q60 88 26 74 Z" fill="#111"/>
      <text x="60" y="40" text-anchor="middle" font-family="Arial Black,Arial" font-size="26" font-weight="900" fill="#fff">REA</text>
      <text x="60" y="102" text-anchor="middle" font-family="Georgia" font-style="italic" font-size="11" font-weight="bold" fill="#F5E100">Desde 1998</text>
      <circle cx="92" cy="82" r="20" fill="#fff" stroke="#111" stroke-width="3"/>
      <polygon points="92,74 99,79 96,88 88,88 85,79" fill="#111"/>
    </svg>"""


def gerar():
    dados = {
        "gerado_em": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "aviso_topo": AVISO_TOPO,
        "ultimo_racha": ULTIMO_RACHA,
        "ranking": dados_ranking(),
        "financeiro": dados_financeiro(),
        "pix": dados_pix(),
        "rsvp_url": RSVP_URL,
        "roster": ROSTER,
        # caixa do REA fica fora do painel por enquanto (só administrativo vê,
        # direto no BALANCETE); para voltar, reincluir "caixa": saldo_caixa()
    }
    html = TEMPLATE.replace("__DADOS__", json.dumps(dados, ensure_ascii=False))
    html = html.replace("__LOGO__", logo_html())
    SAIDA.write_text(html, encoding="utf-8")
    print("Painel gerado:", SAIDA)


TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>REA — Racha Entre Amigos</title>
<meta name="description" content="Ranking do Craque do REA, artilharia, atleta fiel, financeiro e confirmação de presença do racha. Desde 1998.">
<meta name="theme-color" content="#1AA3DD">
<meta property="og:type" content="website">
<meta property="og:title" content="REA — Racha Entre Amigos">
<meta property="og:description" content="Confirme sua presença, veja o ranking, a artilharia e pague a mensalidade por Pix. ⚽">
<meta property="og:image" content="https://rea-1998.github.io/og-image.png">
<meta property="og:url" content="https://rea-1998.github.io/">
<meta name="twitter:card" content="summary_large_image">
<link rel="icon" href="favicon.png" type="image/png">
<link rel="apple-touch-icon" href="icon-180.png">
<link rel="manifest" href="manifest.json">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-title" content="REA">
<style>
  :root{
    --azul:#1AA3DD; --azul-escuro:#0E7CAD; --preto:#141414;
    --amarelo:#F5E100; --branco:#fff; --cinza:#f2f5f7;
    --verde:#1e9e50; --vermelho:#d43c3c;
  }
  *{margin:0;padding:0;box-sizing:border-box;font-family:'Segoe UI',Arial,sans-serif}
  body{background:var(--cinza);color:var(--preto);padding-bottom:40px}
  header{background:linear-gradient(160deg,var(--azul) 0%,var(--azul-escuro) 100%);
    color:#fff;padding:18px 16px 14px;display:flex;align-items:center;gap:14px;
    border-bottom:5px solid var(--preto)}
  .logo{width:74px;height:80px;flex-shrink:0;filter:drop-shadow(0 2px 4px rgba(0,0,0,.35))}
  header h1{font-size:1.35rem;line-height:1.15}
  header .sub{font-size:.8rem;opacity:.92}
  header .desde{color:var(--amarelo);font-weight:700;font-style:italic}
  .ciclo{background:var(--preto);color:var(--amarelo);text-align:center;
    font-size:.78rem;padding:5px;font-weight:600;letter-spacing:.3px}
  main{max-width:760px;margin:0 auto;padding:14px}
  .cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-bottom:14px}
  .card{background:#fff;border-radius:12px;padding:12px 14px;box-shadow:0 1px 4px rgba(0,0,0,.08);
    border-top:4px solid var(--azul)}
  .card .rotulo{font-size:.68rem;text-transform:uppercase;letter-spacing:.6px;color:#777;font-weight:700}
  .card .valor{font-size:1.15rem;font-weight:800;margin-top:3px}
  .card .extra{font-size:.72rem;color:#999}
  .card.ouro{border-top-color:var(--amarelo)}
  nav{display:flex;gap:6px;margin-bottom:12px;flex-wrap:wrap}
  nav button{flex:1;min-width:100px;padding:10px 6px;border:none;border-radius:10px;background:#fff;
    font-weight:700;font-size:.82rem;cursor:pointer;color:#555;box-shadow:0 1px 3px rgba(0,0,0,.08)}
  nav button.ativo{background:var(--azul);color:#fff}
  section{display:none;background:#fff;border-radius:12px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08)}
  section.ativa{display:block}
  h2{font-size:1rem;margin-bottom:10px;color:var(--azul-escuro)}
  table{width:100%;border-collapse:collapse;font-size:.85rem}
  th{text-align:left;font-size:.68rem;text-transform:uppercase;letter-spacing:.5px;color:#888;
    padding:6px 8px;border-bottom:2px solid var(--cinza)}
  td{padding:7px 8px;border-bottom:1px solid var(--cinza)}
  td.num,th.num{text-align:center}
  tr.top1{background:#fff9d6}
  tr.top2{background:#f4f6f8}
  tr.top3{background:#fdf1e3}
  .pos{font-weight:800;color:#999;width:34px}
  .medalha{font-size:1rem}
  .chip{display:inline-block;padding:2px 9px;border-radius:20px;font-size:.7rem;font-weight:700;color:#fff;white-space:nowrap}
  .chip.emdia{background:var(--verde)}
  .chip.atrasado{background:var(--vermelho)}
  .chip.aberto{background:#e08e0b}
  .chip.isento{background:#9aa4ab}
  .fiel{color:var(--verde);font-weight:800}
  .placar{display:flex;flex-wrap:wrap;gap:6px;margin:8px 0}
  .placar span{background:var(--cinza);border-radius:8px;padding:4px 9px;font-size:.75rem;font-weight:600}
  .quadro{display:flex;gap:8px;margin-top:8px;flex-wrap:wrap}
  .qtime{flex:1;min-width:110px;border-radius:10px;padding:8px;text-align:center;font-size:.75rem;font-weight:700}
  .qtime .ved{font-size:.95rem}
  .t-amarelo{background:#fff3a1}.t-azul{background:#c3e5f5}.t-rosa{background:#f8d3dd}
  .destaques{display:flex;gap:10px;margin-top:10px;flex-wrap:wrap}
  .destaques div{flex:1;min-width:130px;border-radius:10px;padding:8px;text-align:center;font-size:.78rem}
  .cheia-box{background:#fff9d6;border:1px solid var(--amarelo)}
  .murcha-box{background:#f6e8e8;border:1px solid #d9b5b5}
  footer{text-align:center;font-size:.7rem;color:#999;margin-top:16px}
  .caixa-pos{color:var(--verde)}.caixa-neg{color:var(--vermelho)}
  .pixbox{display:flex;gap:12px;align-items:center;background:linear-gradient(135deg,#eaf7ef,#f2f5f7);
    border:1px solid #cfe6d8;border-radius:12px;padding:12px;margin-bottom:14px;flex-wrap:wrap}
  .pixtxt{flex:1;min-width:180px}
  .pixtxt b{font-size:.95rem}
  .pixsub{font-size:.72rem;color:#777;margin:2px 0 8px}
  .pixkey{font-family:monospace;font-size:.72rem;background:#fff;border:1px solid #ddd;
    border-radius:6px;padding:6px 8px;word-break:break-all;color:#444}
  .pixbtns{display:flex;gap:6px;margin-top:8px;flex-wrap:wrap}
  .pixbtns button{border:none;background:var(--azul);color:#fff;font-weight:700;font-size:.75rem;
    padding:8px 10px;border-radius:8px;cursor:pointer}
  .pixok{font-size:.72rem;color:var(--verde);font-weight:700;margin-top:6px;min-height:14px}
  .pixqr{width:130px;height:130px;background:#fff;border-radius:8px;padding:4px;flex-shrink:0}
  .rsvp-form{background:var(--cinza);border-radius:12px;padding:14px;margin-bottom:14px}
  #rsvp-nome{width:100%;padding:11px;border-radius:9px;border:1px solid #ccc;font-size:.95rem;background:#fff}
  .rsvp-btns{display:flex;gap:8px;margin-top:10px}
  .rsvp-btns button{flex:1;border:none;border-radius:10px;padding:13px;font-weight:800;font-size:.95rem;cursor:pointer;color:#fff}
  .bt-vou{background:var(--verde)}.bt-nao{background:#9aa4ab}
  .rsvp-msg{margin-top:9px;font-size:.82rem;font-weight:700;min-height:18px;text-align:center}
  .rsvp-msg.ok{color:var(--verde)}.rsvp-msg.erro{color:var(--vermelho)}
  .rsvp-listas{display:flex;gap:10px;flex-wrap:wrap}
  .rsvp-col{flex:1;min-width:140px;background:#fff;border-radius:12px;padding:10px 12px;box-shadow:0 1px 4px rgba(0,0,0,.06)}
  .rsvp-col h3{font-size:.85rem;margin-bottom:6px}
  .cab-vou{color:var(--verde)}.cab-nao{color:#888}.cab-pend{color:#e08e0b}
  .rsvp-col ul{list-style:none}
  .rsvp-col li{padding:5px 0;border-bottom:1px solid var(--cinza);font-size:.88rem}
  .rsvp-col li.vazio{color:#bbb;font-style:italic;border:none}
</style>
</head>
<body>
<header>
  __LOGO__
  <div>
    <h1>REA — Racha Entre Amigos</h1>
    <div class="sub"><span class="desde">Desde 1998</span> · Sábados às 16h30</div>
  </div>
</header>
<div class="ciclo" id="ciclo"></div>
<main>
  <div class="cards" id="cards"></div>

  <nav>
    <button data-aba="presenca" class="ativo">✅ Presença</button>
    <button data-aba="times">👥 Times</button>
    <button data-aba="craque">🏆 Craque</button>
    <button data-aba="artilharia">⚽ Artilharia</button>
    <button data-aba="fiel">⭐ Atleta Fiel</button>
    <button data-aba="estat">📊 Estatísticas</button>
    <button data-aba="fin">💰 Financeiro</button>
    <button data-aba="pagar">💳 Pagar</button>
    <button data-aba="ultimo">📋 Último racha</button>
  </nav>

  <section id="presenca" class="ativa">
    <h2 id="presenca-tit">✅ Confirmar presença</h2>
    <div class="rsvp-form">
      <select id="rsvp-nome"></select>
      <div class="rsvp-btns">
        <button id="rsvp-vou" class="bt-vou">✅ Vou jogar</button>
        <button id="rsvp-nao" class="bt-nao">❌ Não vou</button>
      </div>
      <div class="rsvp-msg" id="rsvp-msg"></div>
    </div>
    <div class="rsvp-listas">
      <div class="rsvp-col">
        <h3 class="cab-vou">✅ Confirmados (<span id="n-vou">0</span>)</h3>
        <ul id="lst-vou"></ul>
      </div>
      <div class="rsvp-col">
        <h3 class="cab-nao">❌ Não vão (<span id="n-nao">0</span>)</h3>
        <ul id="lst-nao"></ul>
      </div>
      <div class="rsvp-col">
        <h3 class="cab-pend">⏳ Ainda não confirmaram (<span id="n-pend">0</span>)</h3>
        <ul id="lst-pend"></ul>
      </div>
    </div>
    <p style="font-size:.72rem;color:#888;margin-top:10px">
      Confirme até as <b>13h de sábado</b>. Quem confirma e falta paga multa de R$ 20.
      Quem está devendo o mês anterior não confirma a partir do 2º sábado do mês.</p>
  </section>

  <section id="times">
    <style>
      .times-grid{display:flex;flex-wrap:wrap;gap:12px}
      .time-card{flex:1 1 240px;background:#fff;border-radius:12px;padding:12px 14px;
                 box-shadow:0 1px 4px rgba(0,0,0,.08);border-top:6px solid #999}
      .tc-azul{border-top-color:#1e88e5}.tc-amarelo{border-top-color:#fdd835}
      .tc-rosa{border-top-color:#ec407a}.tc-vasco{border-top-color:#111}
      .time-card h3{margin:0 0 8px;font-size:1rem}
      .time-card ul{list-style:none;margin:0;padding:0}
      .time-card li{padding:4px 0;border-bottom:1px dashed #eee;display:flex;gap:8px;align-items:center}
      .time-card li:last-child{border-bottom:0}
      .fpts{min-width:26px;text-align:right;font-weight:bold;color:#1a5fb4}
      .fstar{color:#e6a700}
      .pos{font-size:.7rem;color:#888;border:1px solid #ddd;border-radius:6px;padding:0 5px}
      .res-tag{font-size:.65rem;color:#a06000;background:#fff3d6;border-radius:6px;
               padding:1px 6px;margin-left:auto;white-space:nowrap}
      .fiel-li{background:#fff8e1}
    </style>
    <h2 id="times-tit">👥 Times de sábado</h2>
    <div class="times-grid" id="times-grid"></div>
    <p style="font-size:.72rem;color:#888;margin-top:10px">
      Número na frente = <b>pontos de Atleta Fiel</b> (últimos 8 sábados). ⭐ = Atleta Fiel.
      🧤 = goleiro. Quem tem <b>menos pontos</b> no time começa como
      <b>⏳ Reserva (Atleta Fiel)</b> e entra na sequência.</p>
  </section>

  <section id="craque">
    <h2>🏆 Craque do REA 2026</h2>
    <table><thead><tr><th></th><th>Atleta</th><th class="num">Pontos</th></tr></thead><tbody id="tb-craque"></tbody></table>
  </section>

  <section id="artilharia">
    <h2>⚽ Artilharia 2026</h2>
    <table><thead><tr><th></th><th>Atleta</th><th class="num">Gols</th></tr></thead><tbody id="tb-artilharia"></tbody></table>
  </section>

  <section id="fiel">
    <h2>⭐ Atleta Fiel — pontos dos últimos 8 sábados</h2>
    <table><thead><tr><th>Atleta</th><th class="num">Pontos</th><th class="num">Atleta Fiel</th></tr></thead>
      <tbody id="tb-fiel"></tbody></table>
    <p style="font-size:.72rem;color:#888;margin-top:8px">
      Presença vale 2 pontos + 1 de pontualidade (chegar antes das 16h). Quem tem mais pontos tem preferência
      para começar jogando. ⭐ Atleta Fiel = presente nos últimos 8 sábados (mesmo sem jogar) — só é substituído se quiser.</p>
  </section>

  <section id="estat">
    <h2>📊 Estatísticas 2026</h2>
    <table><thead><tr><th>Atleta</th><th class="num">Pts Vit.</th><th class="num">Pres.</th>
      <th class="num">Pontual</th><th class="num">Cartões</th><th class="num">🏅Cheia</th>
      <th class="num">🎈Murcha</th><th class="num">Fiel</th></tr></thead>
      <tbody id="tb-estat"></tbody></table>
    <p style="font-size:.72rem;color:#888;margin-top:8px">
      Atleta Fiel = presente nos últimos 8 sábados (só é substituído se quiser).</p>
  </section>

  <section id="fin">
    <div class="pixbox">
      <div class="pixtxt">
        <b>💳 Quer pagar sua mensalidade?</b>
        <div class="pixsub">Use a aba <b>Pagar</b> — cada atleta tem seu Pix próprio
          e a baixa é <b>automática</b> no painel.</div>
        <div class="pixbtns">
          <button id="btn-ir-pagar">💳 Ir para Pagar</button>
        </div>
      </div>
    </div>
    <h2>💰 Financeiro — mês atual</h2>
    <table><thead><tr><th>Atleta</th><th class="num">Situação</th><th class="num">Saldo (R$)</th></tr></thead>
      <tbody id="tb-fin"></tbody></table>
    <p style="font-size:.72rem;color:#888;margin-top:8px">
      🟢 Em dia · 🟠 Mês em aberto (mensalidade atual, prazo até o fim do mês) · 🔴 Atrasado (deve mês anterior).<br>
      Mensalidade R$ 90 até o fim do mês (R$ 120 em atraso). Quem está devendo não joga a partir do 2º sábado do mês.</p>
  </section>

  <section id="pagar">
    <h2>💳 Pagar minha mensalidade</h2>
    <div class="rsvp-form">
      <select id="pagar-nome"></select>
      <div class="rsvp-msg" id="pagar-msg">Escolha seu nome para ver seu Pix.</div>
    </div>
    <div class="pixbox" id="pagar-card" style="display:none">
      <div class="pixtxt">
        <b>Mensalidade <span id="pagar-mes"></span></b>
        <div class="pixsub">Valor: <b id="pagar-valor"></b></div>
        <div class="pixkey" id="pagar-copia"></div>
        <div class="pixbtns"><button id="btn-copia-pix">📋 Copiar código Pix</button></div>
        <div class="pixok" id="pagar-ok"></div>
      </div>
      <img class="pixqr" id="pagar-qr" alt="QR Pix da sua mensalidade">
    </div>
    <p style="font-size:.72rem;color:#888;margin-top:8px">
      Cada atleta tem um <b>Pix próprio</b> — ao pagar, sua mensalidade recebe baixa
      <b>automática</b> no painel. Mensalidade R$ 90 (R$ 120 em atraso).</p>
  </section>

  <section id="ultimo">
    <h2 id="ultimo-titulo">📋 Último racha</h2>
    <div class="quadro" id="quadro"></div>
    <div class="placar" id="placar"></div>
    <div class="destaques">
      <div class="cheia-box">🏅 <b>Bola Cheia</b><br><span id="cheia"></span></div>
      <div class="murcha-box">🎈 <b>Bola Murcha</b><br><span id="murcha"></span></div>
    </div>
  </section>

  <footer id="rodape"></footer>
</main>
<script>
const D = __DADOS__;
const fmt = n => n.toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2});
document.getElementById('ciclo').textContent = D.aviso_topo;

// cards
const lider = D.ranking.craque[0], art = D.ranking.artilharia[0];
document.getElementById('cards').innerHTML = `
  <div class="card ouro"><div class="rotulo">🏆 Craque do REA</div>
    <div class="valor">${lider.nome}</div><div class="extra">${lider.total} pontos</div></div>
  <div class="card"><div class="rotulo">⚽ Artilheiro</div>
    <div class="valor">${art.nome}</div><div class="extra">${art.total} gols</div></div>
  <div class="card"><div class="rotulo">✅ Atletas fiéis</div>
    <div class="valor">${D.ranking.fieis.length}</div><div class="extra">${D.ranking.fieis.join(', ')||'—'}</div></div>`;

// tabelas de ranking
function tabelaRank(id, itens, unidade){
  const medals = ['🥇','🥈','🥉'];
  document.getElementById(id).innerHTML = itens.map((x,i)=>`
    <tr class="${i<3?'top'+(i+1):''}">
      <td class="pos">${i<3?'<span class=medalha>'+medals[i]+'</span>':(i+1)+'º'}</td>
      <td>${x.nome}</td><td class="num"><b>${x.total}</b></td></tr>`).join('');
}
tabelaRank('tb-craque', D.ranking.craque);
tabelaRank('tb-artilharia', D.ranking.artilharia);

// atleta fiel
document.getElementById('tb-fiel').innerHTML = D.ranking.fiel_tab.map(x=>`
  <tr${x.fiel?' style="background:#e8f6ee"':''}><td>${x.nome}</td>
  <td class="num"><b>${x.pontos}</b></td>
  <td class="num">${x.fiel?'<span class="fiel">⭐ Fiel</span>':''}</td></tr>`).join('');

// estatísticas
document.getElementById('tb-estat').innerHTML = D.ranking.estatisticas.map(x=>`
  <tr><td>${x.nome}</td><td class="num">${x.vitorias}</td><td class="num">${x.presencas}</td>
  <td class="num">${x.pontualidade}</td><td class="num">${x.cartoes||''}</td>
  <td class="num">${x.cheia||''}</td><td class="num">${x.murcha||''}</td>
  <td class="num">${x.fiel?'<span class=fiel>✔</span>':''}</td></tr>`).join('');

// presença ao vivo
const RSVP_URL = D.rsvp_url, ROSTER = D.roster;
const ATRASADOS = new Set(D.financeiro.filter(x=>x.situacao==='Atrasado').map(x=>x.nome));
let sabadoAtual = '';
const selNome = document.getElementById('rsvp-nome');
selNome.innerHTML = '<option value="">— escolha seu nome —</option>' +
  ROSTER.map(n=>`<option>${n}</option>`).join('');
const nomeSalvo = localStorage.getItem('rea_nome');
if (nomeSalvo && ROSTER.includes(nomeSalvo)) selNome.value = nomeSalvo;

function ordinalSab(txt){ return Math.floor((parseInt(txt.split('/')[0],10)-1)/7)+1; }
function rsvpMsg(t,cls){ const e=document.getElementById('rsvp-msg'); e.textContent=t; e.className='rsvp-msg '+(cls||''); }
function pintarLista(d){
  sabadoAtual = d.sabado;
  document.getElementById('presenca-tit').textContent = '✅ Confirmar presença — sábado ' + d.sabado;
  // ordem por posição (goleiro, zagueiro, volante, meia, atacante) e rótulo da posição
  const POSI = D.posicoes || {};
  const POS_ORD = {GOL:0, ZAG:1, VOL:2, MEI:3, ATA:4};
  const porPos = (a,b)=> (POS_ORD[POSI[a]]??9)-(POS_ORD[POSI[b]]??9) || a.localeCompare(b,'pt');
  const li = n => `<li>${n}${POSI[n]?` <span class="pos">${POSI[n]}</span>`:''}</li>`;
  const vou = (d.vou||[]).filter(n=>ROSTER.includes(n)).sort(porPos);
  const nao = (d.naovou||[]).filter(n=>ROSTER.includes(n)).sort(porPos);
  document.getElementById('n-vou').textContent = vou.length;
  document.getElementById('n-nao').textContent = nao.length;
  document.getElementById('lst-vou').innerHTML = vou.length? vou.map(li).join('') : '<li class="vazio">ninguém ainda</li>';
  document.getElementById('lst-nao').innerHTML = nao.length? nao.map(li).join('') : '<li class="vazio">—</li>';
  // quem ainda não respondeu (nem Vou nem Não vou)
  const respondeu = new Set([...vou, ...nao]);
  const pend = ROSTER.filter(n=>!respondeu.has(n)).sort(porPos);
  document.getElementById('n-pend').textContent = pend.length;
  document.getElementById('lst-pend').innerHTML = pend.length? pend.map(li).join('') : '<li class="vazio">todos responderam! 🎉</li>';
}
function carregarPresenca(){
  fetch(RSVP_URL+'?action=list').then(r=>r.json()).then(pintarLista).catch(()=>rsvpMsg('',''));
}
function enviarRSVP(resp){
  const nome = selNome.value;
  if(!nome){ rsvpMsg('Escolha seu nome primeiro 😉','erro'); return; }
  localStorage.setItem('rea_nome', nome);
  if(resp==='Vou' && ATRASADOS.has(nome) && ordinalSab(sabadoAtual)>=2){
    rsvpMsg('⚠️ Você está devendo o mês anterior — acerte o Pix pra poder jogar.','erro'); return;
  }
  rsvpMsg('enviando…','');
  fetch(RSVP_URL+'?action=rsvp&atleta='+encodeURIComponent(nome)+'&resposta='+encodeURIComponent(resp))
    .then(r=>r.json()).then(d=>{ pintarLista(d); rsvpMsg(resp==='Vou'?'✅ Presença confirmada!':'Ok, anotado que não vai.','ok'); })
    .catch(()=>rsvpMsg('Falhou, tenta de novo.','erro'));
}
document.getElementById('rsvp-vou').onclick=()=>enviarRSVP('Vou');
document.getElementById('rsvp-nao').onclick=()=>enviarRSVP('Não vou');
carregarPresenca();

// atalho Financeiro -> aba Pagar
document.getElementById('btn-ir-pagar').onclick = ()=>{
  document.querySelector('nav button[data-aba="pagar"]').click();
  window.scrollTo({top:0, behavior:'smooth'});
};

// financeiro
const CLS_SITUACAO = {'Em dia':'emdia','Atrasado':'atrasado','Mês em aberto':'aberto','Isento':'isento'};
document.getElementById('tb-fin').innerHTML = D.financeiro.map(x=>`
  <tr><td>${x.nome}</td>
    <td class="num"><span class="chip ${CLS_SITUACAO[x.situacao]}">${x.situacao}</span></td>
    <td class="num">${x.situacao==='Isento'?'—':'R$ '+fmt(x.saldo)}</td></tr>`).join('');

// pagar mensalidade (cobrança individual via Pix)
const COB = D.cobrancas || {mes:'', itens:[]};
const cobMap = {}; COB.itens.forEach(i=>cobMap[i.nome]=i);
const selPag = document.getElementById('pagar-nome');
const pagCard = document.getElementById('pagar-card');
const pagMsg = document.getElementById('pagar-msg');
if(!COB.itens.length){
  pagMsg.textContent = 'As cobranças do mês aparecem aqui assim que forem geradas. 😉';
  selPag.style.display = 'none';
} else {
  selPag.innerHTML = '<option value="">— escolha seu nome —</option>' +
    COB.itens.map(i=>`<option>${i.nome}</option>`).join('');
  const salvo = localStorage.getItem('rea_nome');
  if(salvo && cobMap[salvo]) selPag.value = salvo;
}
function mostrarCobranca(){
  const it = cobMap[selPag.value];
  if(!it){ pagCard.style.display='none'; if(COB.itens.length) pagMsg.textContent='Escolha seu nome para ver seu Pix.'; return; }
  localStorage.setItem('rea_nome', selPag.value);
  pagMsg.textContent = '';
  pagCard.style.display = '';
  document.getElementById('pagar-mes').textContent = COB.mes;
  document.getElementById('pagar-valor').textContent = 'R$ ' + it.valor;
  document.getElementById('pagar-copia').textContent = it.copia;
  const qr = document.getElementById('pagar-qr');
  const ok = document.getElementById('pagar-ok');
  const btn = document.getElementById('btn-copia-pix');
  if(it.pago){
    qr.style.display='none'; btn.style.display='none';
    ok.innerHTML = '✅ <b>Pago!</b> Mensalidade em dia.';
  } else {
    qr.style.display=''; qr.src = it.qr; btn.style.display=''; ok.textContent='';
  }
}
if(COB.itens.length){
  selPag.addEventListener('change', mostrarCobranca);
  if(selPag.value) mostrarCobranca();
}
document.getElementById('btn-copia-pix').onclick = ()=>{
  const it = cobMap[selPag.value]; if(!it) return;
  navigator.clipboard.writeText(it.copia).then(()=>{
    const el=document.getElementById('pagar-ok');
    el.textContent='✅ Código Pix copiado! Cole no seu banco.';
    setTimeout(()=>{el.textContent='';},2500);
  });
};

// último racha
const U = D.ultimo_racha;
document.getElementById('ultimo-titulo').textContent = '📋 Último racha — ' + U.data;
document.getElementById('quadro').innerHTML = U.quadro.map(q=>`
  <div class="qtime t-${q.time.toLowerCase()}">${q.time}<br>
  <span class="ved">${q.v}V ${q.e}E ${q.d}D</span><br>${q.gols} gols</div>`).join('');
document.getElementById('placar').innerHTML = U.partidas.map((p,i)=>`<span>${i+1}· ${p}</span>`).join('');
document.getElementById('cheia').textContent = U.bola_cheia;
document.getElementById('murcha').textContent = U.bola_murcha;

// times de sábado (montados pelo bot)
const TM = D.times || {sabado:'', times:[]};
if(TM.times.length){
  document.getElementById('times-tit').textContent = '👥 Times — sábado ' + TM.sabado;
  document.getElementById('times-grid').innerHTML = TM.times.map(t=>{
    const row = a => `<li class="${a.fiel?'fiel-li':''}"><span class="fpts">${a.fiel_pts}</span>`+
      `<span class="fstar">${a.fiel?'⭐':''}</span>`+
      `<span>${a.goleiro?'🧤 ':''}${a.nome}</span><span class="pos">${a.pos}</span>`+
      `${a.titular?'':'<span class="res-tag">⏳ Reserva (Atleta Fiel)</span>'}</li>`;
    return `<div class="time-card tc-${t.cor.toLowerCase()}"><h3>${t.cor.toUpperCase()}</h3>`+
      `<ul>${t.atletas.map(row).join('')}</ul></div>`;
  }).join('');
} else {
  document.getElementById('times-grid').innerHTML =
    '<p style="color:#888">Os times aparecem aqui assim que forem montados (normalmente no sábado, após as confirmações). 😉</p>';
}

// abas
document.querySelectorAll('nav button').forEach(b=>b.onclick=()=>{
  document.querySelectorAll('nav button').forEach(x=>x.classList.remove('ativo'));
  document.querySelectorAll('section').forEach(x=>x.classList.remove('ativa'));
  b.classList.add('ativo');
  document.getElementById(b.dataset.aba).classList.add('ativa');
});

document.getElementById('rodape').textContent =
  'Atualizado em ' + D.gerado_em + ' · Tesouraria: Mateus';
</script>
</body>
</html>
"""

if __name__ == "__main__":
    gerar()
